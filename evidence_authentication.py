"""
Digital Evidence Authentication and Chain of Custody System
Ensures court admissibility through proper evidence handling procedures
"""

import hashlib
import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging
from cryptography.hazmat.primitives import serialization, hashes
from cryptography.hazmat.primitives.asymmetric import rsa, padding
from cryptography.hazmat.primitives.serialization import Encoding, PrivateFormat, NoEncryption
import base64

logger = logging.getLogger(__name__)


class DigitalEvidenceAuthenticator:
    """Handles digital evidence authentication and chain of custody"""

    def __init__(self, evidence_dir: Path = None):
        self.evidence_dir = evidence_dir or Path("evidence")
        self.evidence_dir.mkdir(exist_ok=True)
        self.chain_of_custody_file = self.evidence_dir / "chain_of_custody.json"
        self.private_key = self._load_or_generate_signing_key()

    def _load_or_generate_signing_key(self) -> rsa.RSAPrivateKey:
        """Load or generate RSA key for digital signing"""
        key_file = self.evidence_dir / "signing_key.pem"

        if key_file.exists():
            with open(key_file, 'rb') as f:
                private_key = serialization.load_pem_private_key(
                    f.read(), password=None
                )
        else:
            # Generate new key pair
            private_key = rsa.generate_private_key(
                public_exponent=65537,
                key_size=2048
            )

            # Save private key
            with open(key_file, 'wb') as f:
                f.write(private_key.private_bytes(
                    encoding=Encoding.PEM,
                    format=PrivateFormat.PKCS8,
                    encryption_algorithm=NoEncryption()
                ))

            # Save public key
            public_key = private_key.public_key()
            with open(self.evidence_dir / "public_key.pem", 'wb') as f:
                f.write(public_key.public_bytes(
                    encoding=Encoding.PEM,
                    format=serialization.PublicFormat.SubjectPublicKeyInfo
                ))

        return private_key

    def authenticate_document(self, file_path: Path, custodian: str,
                            collection_method: str = "automated") -> Dict[str, Any]:
        """Create comprehensive authentication record for a document"""

        # Calculate multiple hashes for integrity verification
        file_hashes = self._calculate_file_hashes(file_path)

        # Get file metadata
        metadata = self._extract_file_metadata(file_path)

        # Create authentication record
        auth_record = {
            "file_path": str(file_path),
            "file_name": file_path.name,
            "authenticated_at": datetime.now(timezone.utc).isoformat(),
            "custodian": custodian,
            "collection_method": collection_method,
            "file_size": file_path.stat().st_size,
            "modification_time": datetime.fromtimestamp(
                file_path.stat().st_mtime, timezone.utc
            ).isoformat(),
            "creation_time": datetime.fromtimestamp(
                file_path.stat().st_ctime, timezone.utc
            ).isoformat(),
            "hashes": file_hashes,
            "metadata": metadata,
            "authentication_method": "digital_signature",
            "compliance": {
                "fed_rules_evidence": ["901", "902", "1001"],
                "best_evidence_rule": True,
                "hearsay_exceptions": ["803(6)", "902(11)", "902(13)"]
            }
        }

        # Add digital signature
        signature = self._sign_record(auth_record)
        auth_record["digital_signature"] = signature
        auth_record["signature_algorithm"] = "RSA-2048-SHA256"

        # Add to chain of custody
        self._add_to_chain_of_custody(auth_record)

        return auth_record

    def _calculate_file_hashes(self, file_path: Path) -> Dict[str, str]:
        """Calculate multiple hash algorithms for file integrity"""
        hashes_dict = {}

        # Hash algorithms for different purposes
        algorithms = {
            "md5": hashlib.md5(),
            "sha1": hashlib.sha1(),
            "sha256": hashlib.sha256(),
            "sha512": hashlib.sha512()
        }

        with open(file_path, 'rb') as f:
            # Read file in chunks to handle large files
            while chunk := f.read(8192):
                for algo in algorithms.values():
                    algo.update(chunk)

        for name, algo in algorithms.items():
            hashes_dict[name] = algo.hexdigest()

        return hashes_dict

    def _extract_file_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Extract technical metadata for authentication"""
        metadata = {
            "file_type": file_path.suffix.lower(),
            "permissions": oct(file_path.stat().st_mode)[-3:],
            "inode": file_path.stat().st_ino if hasattr(file_path.stat(), 'st_ino') else None,
        }

        # Add file-type specific metadata
        if file_path.suffix.lower() == '.pdf':
            metadata.update(self._extract_pdf_metadata(file_path))
        elif file_path.suffix.lower() in ['.xlsx', '.xls']:
            metadata.update(self._extract_excel_metadata(file_path))
        elif file_path.suffix.lower() in ['.jpg', '.jpeg', '.png']:
            metadata.update(self._extract_image_metadata(file_path))

        return metadata

    def _extract_pdf_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Extract PDF-specific metadata"""
        try:
            import PyPDF2
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                metadata = reader.metadata or {}
                return {
                    "pdf_version": reader.pdf_header if hasattr(reader, 'pdf_header') else None,
                    "page_count": len(reader.pages),
                    "creator": metadata.get('/Creator', ''),
                    "producer": metadata.get('/Producer', ''),
                    "creation_date": str(metadata.get('/CreationDate', '')),
                    "modification_date": str(metadata.get('/ModDate', '')),
                    "encrypted": reader.is_encrypted
                }
        except Exception as e:
            logger.warning(f"Could not extract PDF metadata: {e}")
            return {"pdf_extraction_error": str(e)}

    def _extract_excel_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Extract Excel-specific metadata"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            props = workbook.properties
            return {
                "creator": props.creator,
                "last_modified_by": props.lastModifiedBy,
                "created": props.created.isoformat() if props.created else None,
                "modified": props.modified.isoformat() if props.modified else None,
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames
            }
        except Exception as e:
            logger.warning(f"Could not extract Excel metadata: {e}")
            return {"excel_extraction_error": str(e)}

    def _extract_image_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Extract image EXIF metadata"""
        try:
            from PIL import Image
            from PIL.ExifTags import TAGS

            with Image.open(file_path) as img:
                exif_data = img.getexif()
                metadata = {
                    "image_format": img.format,
                    "image_mode": img.mode,
                    "image_size": img.size
                }

                if exif_data:
                    exif_dict = {}
                    for tag_id, value in exif_data.items():
                        tag = TAGS.get(tag_id, tag_id)
                        exif_dict[tag] = str(value)
                    metadata["exif"] = exif_dict

                return metadata
        except Exception as e:
            logger.warning(f"Could not extract image metadata: {e}")
            return {"image_extraction_error": str(e)}

    def _sign_record(self, record: Dict[str, Any]) -> str:
        """Digitally sign the authentication record"""
        # Create canonical representation
        record_copy = record.copy()
        if "digital_signature" in record_copy:
            del record_copy["digital_signature"]
        if "signature_algorithm" in record_copy:
            del record_copy["signature_algorithm"]

        canonical_data = json.dumps(record_copy, sort_keys=True, separators=(',', ':'))

        # Sign the data
        signature = self.private_key.sign(
            canonical_data.encode('utf-8'),
            padding.PSS(
                mgf=padding.MGF1(hashes.SHA256()),
                salt_length=padding.PSS.MAX_LENGTH
            ),
            hashes.SHA256()
        )

        return base64.b64encode(signature).decode('utf-8')

    def verify_signature(self, record: Dict[str, Any]) -> bool:
        """Verify digital signature of authentication record"""
        try:
            signature = base64.b64decode(record["digital_signature"])

            # Recreate canonical data
            record_copy = record.copy()
            del record_copy["digital_signature"]
            del record_copy["signature_algorithm"]
            canonical_data = json.dumps(record_copy, sort_keys=True, separators=(',', ':'))

            # Verify signature
            public_key = self.private_key.public_key()
            public_key.verify(
                signature,
                canonical_data.encode('utf-8'),
                padding.PSS(
                    mgf=padding.MGF1(hashes.SHA256()),
                    salt_length=padding.PSS.MAX_LENGTH
                ),
                hashes.SHA256()
            )
            return True
        except Exception as e:
            logger.error(f"Signature verification failed: {e}")
            return False

    def _add_to_chain_of_custody(self, auth_record: Dict[str, Any]):
        """Add authentication record to chain of custody log"""
        chain_entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "action": "document_authenticated",
            "file_name": auth_record["file_name"],
            "custodian": auth_record["custodian"],
            "hash_sha256": auth_record["hashes"]["sha256"],
            "authentication_id": self._generate_auth_id(auth_record)
        }

        # Load existing chain
        chain_of_custody = []
        if self.chain_of_custody_file.exists():
            with open(self.chain_of_custody_file, 'r') as f:
                chain_of_custody = json.load(f)

        # Add new entry
        chain_of_custody.append(chain_entry)

        # Save updated chain
        with open(self.chain_of_custody_file, 'w') as f:
            json.dump(chain_of_custody, f, indent=2)

    def _generate_auth_id(self, auth_record: Dict[str, Any]) -> str:
        """Generate unique authentication ID"""
        data = f"{auth_record['file_name']}{auth_record['authenticated_at']}{auth_record['custodian']}"
        return hashlib.sha256(data.encode()).hexdigest()[:16]

    def create_custody_transfer_record(self, file_name: str, from_custodian: str,
                                     to_custodian: str, reason: str) -> Dict[str, Any]:
        """Record transfer of custody"""
        transfer_record = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "action": "custody_transfer",
            "file_name": file_name,
            "from_custodian": from_custodian,
            "to_custodian": to_custodian,
            "reason": reason,
            "transfer_id": hashlib.sha256(
                f"{file_name}{from_custodian}{to_custodian}{datetime.now().isoformat()}".encode()
            ).hexdigest()[:16]
        }

        # Add to chain of custody
        chain_of_custody = []
        if self.chain_of_custody_file.exists():
            with open(self.chain_of_custody_file, 'r') as f:
                chain_of_custody = json.load(f)

        chain_of_custody.append(transfer_record)

        with open(self.chain_of_custody_file, 'w') as f:
            json.dump(chain_of_custody, f, indent=2)

        return transfer_record

    def verify_file_integrity(self, file_path: Path, original_record: Dict[str, Any]) -> Dict[str, Any]:
        """Verify file has not been altered since authentication"""
        current_hashes = self._calculate_file_hashes(file_path)
        original_hashes = original_record["hashes"]

        integrity_check = {
            "file_path": str(file_path),
            "verified_at": datetime.now(timezone.utc).isoformat(),
            "integrity_status": "verified" if current_hashes == original_hashes else "compromised",
            "hash_comparison": {}
        }

        for algo in current_hashes:
            integrity_check["hash_comparison"][algo] = {
                "original": original_hashes.get(algo),
                "current": current_hashes[algo],
                "matches": current_hashes[algo] == original_hashes.get(algo)
            }

        return integrity_check

    def generate_custody_affidavit(self, file_names: List[str], affiant: str,
                                 case_number: str) -> str:
        """Generate chain of custody affidavit for court"""

        # Get chain of custody records
        chain_records = []
        if self.chain_of_custody_file.exists():
            with open(self.chain_of_custody_file, 'r') as f:
                all_records = json.load(f)
                chain_records = [r for r in all_records if r["file_name"] in file_names]

        affidavit = f"""
AFFIDAVIT OF CHAIN OF CUSTODY AND DIGITAL EVIDENCE AUTHENTICATION

STATE OF ILLINOIS       )
                        ) SS.
COUNTY OF COOK          )

I, {affiant}, being first duly sworn, depose and state:

1. I am competent to testify to the matters contained herein based upon my personal knowledge and experience.

2. I am responsible for the collection, preservation, and authentication of digital evidence in Case No. {case_number}.

3. The following digital evidence has been properly collected, authenticated, and maintained in my custody:

"""

        for i, file_name in enumerate(file_names, 1):
            file_records = [r for r in chain_records if r["file_name"] == file_name]
            if file_records:
                auth_record = file_records[0]
                affidavit += f"""
   {i}. File: {file_name}
      - Authenticated on: {auth_record.get('timestamp', 'N/A')}
      - SHA-256 Hash: {auth_record.get('hash_sha256', 'N/A')}
      - Authentication ID: {auth_record.get('authentication_id', 'N/A')}
"""

        affidavit += f"""

4. Each digital file has been authenticated using cryptographic hash functions (MD5, SHA-1, SHA-256, SHA-512) to ensure integrity.

5. All files have been digitally signed using RSA-2048 encryption with SHA-256 hashing to prevent tampering.

6. The chain of custody has been maintained continuously, and all transfers have been properly documented.

7. The authentication methods employed comply with Federal Rules of Evidence 901, 902, and 1001-1008.

8. These digital files are exact duplicates of the original evidence and have not been altered in any way.

9. The metadata and technical specifications of each file have been preserved and documented.

10. I certify that the attached digital evidence is authentic, reliable, and admissible under the Federal Rules of Evidence.

Further affiant sayeth not.

_________________________________
{affiant}
Digital Evidence Custodian

Subscribed and sworn to before me this _____ day of _____________, 2024.

_________________________________
Notary Public

My commission expires: _____________

AUTHENTICATION CERTIFICATES ATTACHED:
- Digital signatures and hash verification
- Technical metadata reports
- Chain of custody log
- Integrity verification results
"""

        return affidavit

    def export_authentication_package(self, file_names: List[str],
                                    output_dir: Path) -> Dict[str, Any]:
        """Export complete authentication package for court submission"""
        output_dir.mkdir(exist_ok=True)

        # Load all authentication records
        auth_records = {}
        integrity_checks = {}

        for file_name in file_names:
            # Find authentication record
            if self.chain_of_custody_file.exists():
                with open(self.chain_of_custody_file, 'r') as f:
                    records = json.load(f)
                    for record in records:
                        if record.get("file_name") == file_name and "hash_sha256" in record:
                            # Load full authentication record
                            auth_file = self.evidence_dir / f"{record.get('authentication_id', file_name)}_auth.json"
                            if auth_file.exists():
                                with open(auth_file, 'r') as af:
                                    auth_records[file_name] = json.load(af)
                            break

            # Verify current integrity
            file_path = Path(file_name)
            if file_path.exists() and file_name in auth_records:
                integrity_checks[file_name] = self.verify_file_integrity(
                    file_path, auth_records[file_name]
                )

        # Export files
        package_files = {
            "authentication_records.json": auth_records,
            "integrity_verification.json": integrity_checks,
            "chain_of_custody.json": chain_records if self.chain_of_custody_file.exists() else [],
            "public_key.pem": (self.evidence_dir / "public_key.pem").read_text() if (self.evidence_dir / "public_key.pem").exists() else None
        }

        for filename, content in package_files.items():
            if content is not None:
                output_file = output_dir / filename
                if filename.endswith('.json'):
                    with open(output_file, 'w') as f:
                        json.dump(content, f, indent=2)
                else:
                    with open(output_file, 'w') as f:
                        f.write(content)

        return {
            "package_location": str(output_dir),
            "files_authenticated": len(auth_records),
            "integrity_verified": sum(1 for check in integrity_checks.values()
                                   if check["integrity_status"] == "verified"),
            "package_files": list(package_files.keys())
        }